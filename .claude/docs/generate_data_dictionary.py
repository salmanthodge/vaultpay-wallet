"""
VaultPay — Data Dictionary generator.

Single source of truth for the cross-service table catalog. Re-run after adding
or changing tables in any service:  py docs/generate_data_dictionary.py

Output: docs/vaultpay-data-dictionary.xlsx
  - "Tables"   : one row per table (all services) — Service / Table / Purpose / Reason / Description
  - one sheet per service : column-level dictionary (Table / Column / Type / Constraints / Description)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# DATA  (extend this as each service is built — keep it in sync with prisma)
# ---------------------------------------------------------------------------

# Table-level catalog: (service, table, purpose, reason/why, description)
TABLES = [
    # ---- auth-service -----------------------------------------------------
    ("auth", "users",
     "Stores customer accounts.",
     "Customers are the primary actors; auth-service owns the identity record so other services reference users by id only.",
     "Core identity row: credentials, contact, status and verification/login metadata for a customer."),
    ("auth", "roles",
     "Defines named roles for RBAC.",
     "auth-service is the RBAC source of truth; roles are defined once and consulted by every service.",
     "Role registry (e.g. CUSTOMER, ADMIN). is_system marks built-in roles that must not be deleted."),
    ("auth", "permissions",
     "Defines granular permissions.",
     "Permissions are checked by the rbac middleware; centralizing them keeps authorization consistent.",
     "Permission registry of action codes (e.g. WALLET_READ, DOCUMENT_DELETE)."),
    ("auth", "role_permissions",
     "Maps roles to permissions.",
     "Many-to-many between roles and permissions; lets roles be composed from reusable permissions.",
     "Join table linking a role to each permission it grants."),
    ("auth", "user_roles",
     "Maps users to roles.",
     "A user may hold multiple roles; a join table keeps assignments flexible and auditable.",
     "Join table linking a user to each assigned role."),
    ("auth", "refresh_tokens",
     "Durable record of issued refresh tokens.",
     "Enables refresh-token rotation with reuse detection and audit; Redis holds the hot allow/deny path, this table is the durable family record.",
     "Hashed refresh tokens grouped by rotation family, with expiry, revocation and replacement chain."),
    ("auth", "email_verifications",
     "Pending email-verification tokens.",
     "Email ownership must be proven before an account is fully active; tokens are short-lived and single-use.",
     "Hashed verification tokens with expiry and consumed_at; one active token per verification request."),
    ("auth", "password_resets",
     "Pending password-reset tokens.",
     "Self-service password recovery needs secure, expiring, single-use tokens delivered by email.",
     "Hashed reset tokens with expiry and consumed_at, tied to a user."),
    ("auth", "admin_users",
     "Stores back-office admin accounts.",
     "Admins authenticate via sessions (not JWT) and are kept separate from customers for blast-radius isolation.",
     "Admin identity row with role enum and status; session lives in Redis, this is the durable account."),
    ("auth", "service_clients",
     "Registered services allowed to request S2S tokens.",
     "Service-to-service calls must be authenticated; each caller is a registered client with a hashed secret and scopes.",
     "S2S client registry: client_id, hashed secret, allowed scopes and active flag."),
    ("auth", "auth_events",
     "Audit log of authentication events.",
     "VAPT/security and incident response require a durable trail of logins, failures, refreshes and resets with IP/geo.",
     "Append-only audit rows: actor, event type, IP, geo, user-agent and metadata."),
    ("auth", "oauth_accounts",
     "Links a user to an external OAuth identity.",
     "Social/OAuth login (Google, GitHub) must map a provider account to one VaultPay user without duplicating identities.",
     "One row per linked provider account: provider, provider subject id, email and (encrypted) provider tokens."),
    ("auth", "user_mfa",
     "Per-user multi-factor (TOTP) configuration.",
     "MFA hardens login; the TOTP secret and enabled state must persist per user (one factor row per user).",
     "Encrypted TOTP secret with enabled flag and timestamps; backup codes live in mfa_backup_codes."),
    ("auth", "mfa_backup_codes",
     "Single-use MFA recovery codes.",
     "Users locked out of their authenticator need hashed, single-use recovery codes to regain access.",
     "Hashed backup codes per user with consumed_at to enforce single use."),
    ("auth", "user_devices",
     "Known customer devices / sessions.",
     "Device tracking enables trusted-device skips, session listing and remote revocation, and feeds security audit.",
     "Per-device record: fingerprint, name, IP, geo, user-agent, trust flag, last-seen and revocation."),
    # ---- wallet-service ------------------------------------------------
    ("wallet", "wallets",
     "Stores a customer's balance per currency.",
     "Each customer can hold one wallet per currency; wallet-service owns balances so other services never touch money directly.",
     "Balance (Decimal) with status and currency, keyed by user_id (plain reference to auth users.id). Unique per (user_id, currency)."),
    ("wallet", "transactions",
     "Append-only ledger of balance movements.",
     "Every credit/debit must be auditable and idempotent; the ledger is the source of truth for balance history.",
     "One row per balance movement: type (CREDIT/DEBIT), amount, balance_after, status, optional idempotency reference and linked transfer."),
    ("wallet", "transfers",
     "Records wallet-to-wallet transfers.",
     "A transfer is a single logical operation spanning two ledger entries; it needs its own status and idempotency key.",
     "From/to wallets, amount, currency, status (PENDING/COMPLETED/FAILED/REVERSED) and a unique reference; links the two transaction legs."),
    # ---- vault-service -------------------------------------------------
    ("vault", "documents",
     "Metadata for an uploaded document.",
     "The encrypted file lives in MinIO; the service needs a queryable record of ownership, name, size, checksum and status.",
     "Per-document row: user_id (plain ref), original filename, MinIO storage_key, mime type, plaintext size, sha256 checksum, category and soft-delete status."),
    ("vault", "document_shares",
     "Time-limited public share links for a document.",
     "Owners can share a document without exposing their JWT; tokens must be revocable, expiring and download-limited.",
     "Hashed share token with optional expiry and max-downloads, a download counter, creator and revocation timestamp."),
]

# Column-level dictionary per service:
#   service -> list of (table, column, type, constraints, description)
COLUMNS = {
    "auth": [
        # users
        ("users", "id", "uuid", "PK, default uuid", "Primary key."),
        ("users", "email", "varchar", "UNIQUE, NOT NULL", "Login identifier; stored lowercased."),
        ("users", "password_hash", "varchar", "NULL", "bcrypt hash; null for OAuth-only accounts; never plaintext."),
        ("users", "full_name", "varchar", "NULL", "Customer display name."),
        ("users", "phone", "varchar", "UNIQUE, NULL", "Optional contact / future MFA."),
        ("users", "status", "enum", "default PENDING", "PENDING | ACTIVE | SUSPENDED | DELETED."),
        ("users", "email_verified_at", "timestamptz", "NULL", "Set when email verification succeeds."),
        ("users", "last_login_at", "timestamptz", "NULL", "Last successful login timestamp."),
        ("users", "failed_login_count", "int", "default 0", "Reset on success; drives lockout."),
        ("users", "created_at", "timestamptz", "default now()", "Row creation time."),
        ("users", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        ("users", "deleted_at", "timestamptz", "NULL", "Soft-delete marker."),
        # roles
        ("roles", "id", "uuid", "PK", "Primary key."),
        ("roles", "name", "varchar", "UNIQUE, NOT NULL", "Role name (e.g. CUSTOMER, ADMIN)."),
        ("roles", "description", "varchar", "NULL", "Human description of the role."),
        ("roles", "is_system", "boolean", "default false", "Built-in role; protected from deletion."),
        ("roles", "created_at", "timestamptz", "default now()", "Row creation time."),
        ("roles", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # permissions
        ("permissions", "id", "uuid", "PK", "Primary key."),
        ("permissions", "code", "varchar", "UNIQUE, NOT NULL", "Permission code (e.g. WALLET_READ)."),
        ("permissions", "description", "varchar", "NULL", "What the permission allows."),
        ("permissions", "created_at", "timestamptz", "default now()", "Row creation time."),
        ("permissions", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # role_permissions
        ("role_permissions", "id", "uuid", "PK", "Primary key."),
        ("role_permissions", "role_id", "uuid", "FK roles.id", "Owning role."),
        ("role_permissions", "permission_id", "uuid", "FK permissions.id", "Granted permission."),
        ("role_permissions", "(role_id, permission_id)", "-", "UNIQUE", "Prevents duplicate grants."),
        # user_roles
        ("user_roles", "id", "uuid", "PK", "Primary key."),
        ("user_roles", "user_id", "uuid", "FK users.id", "User receiving the role."),
        ("user_roles", "role_id", "uuid", "FK roles.id", "Assigned role."),
        ("user_roles", "(user_id, role_id)", "-", "UNIQUE", "Prevents duplicate assignment."),
        ("user_roles", "assigned_at", "timestamptz", "default now()", "When the role was granted."),
        # refresh_tokens
        ("refresh_tokens", "id", "uuid", "PK", "Primary key."),
        ("refresh_tokens", "user_id", "uuid", "FK users.id", "Token owner."),
        ("refresh_tokens", "token_hash", "varchar", "NOT NULL, indexed", "Hash of the refresh token."),
        ("refresh_tokens", "family_id", "uuid", "NOT NULL", "Rotation family for reuse detection."),
        ("refresh_tokens", "expires_at", "timestamptz", "NOT NULL", "Expiry of this token."),
        ("refresh_tokens", "revoked_at", "timestamptz", "NULL", "Set on logout/rotation/abuse."),
        ("refresh_tokens", "replaced_by", "uuid", "NULL", "Successor token id after rotation."),
        ("refresh_tokens", "ip", "varchar", "NULL", "Issuing client IP."),
        ("refresh_tokens", "user_agent", "varchar", "NULL", "Issuing client user-agent."),
        ("refresh_tokens", "user_device_id", "uuid", "FK user_devices.id, NULL", "Device this token belongs to (enables remote revocation)."),
        ("refresh_tokens", "created_at", "timestamptz", "default now()", "Issue time."),
        # email_verifications
        ("email_verifications", "id", "uuid", "PK", "Primary key."),
        ("email_verifications", "user_id", "uuid", "FK users.id", "User being verified."),
        ("email_verifications", "token_hash", "varchar", "NOT NULL", "Hash of the verification token."),
        ("email_verifications", "expires_at", "timestamptz", "NOT NULL", "Token expiry."),
        ("email_verifications", "consumed_at", "timestamptz", "NULL", "Set when used (single-use)."),
        ("email_verifications", "created_at", "timestamptz", "default now()", "Creation time."),
        # password_resets
        ("password_resets", "id", "uuid", "PK", "Primary key."),
        ("password_resets", "user_id", "uuid", "FK users.id", "User resetting password."),
        ("password_resets", "token_hash", "varchar", "NOT NULL", "Hash of the reset token."),
        ("password_resets", "expires_at", "timestamptz", "NOT NULL", "Token expiry."),
        ("password_resets", "consumed_at", "timestamptz", "NULL", "Set when used (single-use)."),
        ("password_resets", "created_at", "timestamptz", "default now()", "Creation time."),
        # admin_users
        ("admin_users", "id", "uuid", "PK", "Primary key."),
        ("admin_users", "email", "varchar", "UNIQUE, NOT NULL", "Admin login email."),
        ("admin_users", "password_hash", "varchar", "NOT NULL", "Argon2/bcrypt hash."),
        ("admin_users", "full_name", "varchar", "NULL", "Admin display name."),
        ("admin_users", "role", "enum", "NOT NULL", "SUPER_ADMIN | ADMIN | SUPPORT | AUDITOR."),
        ("admin_users", "status", "enum", "default ACTIVE", "ACTIVE | SUSPENDED."),
        ("admin_users", "last_login_at", "timestamptz", "NULL", "Last successful admin login."),
        ("admin_users", "created_at", "timestamptz", "default now()", "Creation time."),
        ("admin_users", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        ("admin_users", "deleted_at", "timestamptz", "NULL", "Soft-delete marker."),
        # service_clients
        ("service_clients", "id", "uuid", "PK", "Primary key."),
        ("service_clients", "client_id", "varchar", "UNIQUE, NOT NULL", "Public client identifier."),
        ("service_clients", "client_secret_hash", "varchar", "NOT NULL", "Hashed client secret."),
        ("service_clients", "name", "varchar", "NOT NULL", "Service name (e.g. wallet-service)."),
        ("service_clients", "scopes", "jsonb", "default []", "Allowed S2S scopes."),
        ("service_clients", "is_active", "boolean", "default true", "Disables the client when false."),
        ("service_clients", "created_at", "timestamptz", "default now()", "Creation time."),
        ("service_clients", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # auth_events
        ("auth_events", "id", "uuid", "PK", "Primary key."),
        ("auth_events", "actor_type", "enum", "NOT NULL", "CUSTOMER | ADMIN | SERVICE."),
        ("auth_events", "actor_id", "uuid", "NULL", "User/admin/client id when known."),
        ("auth_events", "event_type", "varchar", "NOT NULL", "LOGIN_SUCCESS, LOGIN_FAILED, LOGOUT, TOKEN_REFRESH, PASSWORD_RESET..."),
        ("auth_events", "ip", "varchar", "NULL", "Client IP (from ipTracker)."),
        ("auth_events", "geo_country", "varchar", "NULL", "Geo country (geoip-lite)."),
        ("auth_events", "user_agent", "varchar", "NULL", "Client user-agent."),
        ("auth_events", "metadata", "jsonb", "NULL", "Extra event context."),
        ("auth_events", "created_at", "timestamptz", "default now()", "Event time."),
        # oauth_accounts
        ("oauth_accounts", "id", "uuid", "PK", "Primary key."),
        ("oauth_accounts", "user_id", "uuid", "FK users.id", "Linked VaultPay user."),
        ("oauth_accounts", "provider", "enum", "NOT NULL", "GOOGLE | GITHUB | ... (free providers)."),
        ("oauth_accounts", "provider_user_id", "varchar", "NOT NULL", "Provider subject (sub) id."),
        ("oauth_accounts", "email", "varchar", "NULL", "Email reported by the provider."),
        ("oauth_accounts", "access_token_enc", "varchar", "NULL", "Encrypted provider access token (optional)."),
        ("oauth_accounts", "refresh_token_enc", "varchar", "NULL", "Encrypted provider refresh token (optional)."),
        ("oauth_accounts", "(provider, provider_user_id)", "-", "UNIQUE", "One VaultPay link per provider identity."),
        ("oauth_accounts", "linked_at", "timestamptz", "default now()", "When the account was linked."),
        ("oauth_accounts", "created_at", "timestamptz", "default now()", "Creation time."),
        ("oauth_accounts", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # user_mfa
        ("user_mfa", "id", "uuid", "PK", "Primary key."),
        ("user_mfa", "user_id", "uuid", "FK users.id, UNIQUE", "Owner (one MFA config per user)."),
        ("user_mfa", "type", "enum", "default TOTP", "Factor type (TOTP for now)."),
        ("user_mfa", "secret_enc", "varchar", "NOT NULL", "Encrypted TOTP shared secret."),
        ("user_mfa", "is_enabled", "boolean", "default false", "True once verified and active."),
        ("user_mfa", "enabled_at", "timestamptz", "NULL", "When MFA was enabled."),
        ("user_mfa", "created_at", "timestamptz", "default now()", "Creation time."),
        ("user_mfa", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # mfa_backup_codes
        ("mfa_backup_codes", "id", "uuid", "PK", "Primary key."),
        ("mfa_backup_codes", "user_id", "uuid", "FK users.id", "Owner of the code."),
        ("mfa_backup_codes", "code_hash", "varchar", "NOT NULL", "Hash of a single backup code."),
        ("mfa_backup_codes", "consumed_at", "timestamptz", "NULL", "Set when used (single-use)."),
        ("mfa_backup_codes", "created_at", "timestamptz", "default now()", "Creation time."),
        # user_devices
        ("user_devices", "id", "uuid", "PK", "Primary key."),
        ("user_devices", "user_id", "uuid", "FK users.id", "Device owner."),
        ("user_devices", "device_id", "varchar", "NOT NULL", "Client fingerprint / device id."),
        ("user_devices", "name", "varchar", "NULL", "Friendly device label."),
        ("user_devices", "ip", "varchar", "NULL", "Last-seen IP."),
        ("user_devices", "user_agent", "varchar", "NULL", "Last-seen user-agent."),
        ("user_devices", "geo_country", "varchar", "NULL", "Geo country (geoip-lite)."),
        ("user_devices", "is_trusted", "boolean", "default false", "Skip MFA on trusted device."),
        ("user_devices", "last_seen_at", "timestamptz", "default now()", "Last activity time."),
        ("user_devices", "revoked_at", "timestamptz", "NULL", "Set on remote sign-out."),
        ("user_devices", "created_at", "timestamptz", "default now()", "First seen time."),
    ],
    "wallet": [
        # wallets
        ("wallets", "id", "uuid", "PK", "Primary key."),
        ("wallets", "user_id", "uuid", "indexed", "Owner — plain reference to auth users.id (no cross-DB FK)."),
        ("wallets", "currency", "varchar(3)", "NOT NULL", "ISO 4217 currency code."),
        ("wallets", "balance", "decimal(20,4)", "default 0", "Current balance; updated atomically."),
        ("wallets", "status", "enum", "default ACTIVE", "ACTIVE | FROZEN | CLOSED."),
        ("wallets", "(user_id, currency)", "-", "UNIQUE", "One wallet per currency per user."),
        ("wallets", "created_at", "timestamptz", "default now()", "Creation time."),
        ("wallets", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        # transactions
        ("transactions", "id", "uuid", "PK", "Primary key."),
        ("transactions", "wallet_id", "uuid", "FK wallets.id", "Wallet the entry belongs to."),
        ("transactions", "type", "enum", "NOT NULL", "CREDIT | DEBIT."),
        ("transactions", "amount", "decimal(20,4)", "NOT NULL", "Movement amount (positive)."),
        ("transactions", "balance_after", "decimal(20,4)", "NOT NULL", "Wallet balance after this entry."),
        ("transactions", "status", "enum", "default COMPLETED", "PENDING | COMPLETED | FAILED."),
        ("transactions", "reference", "varchar", "UNIQUE, NULL", "Idempotency key (optional)."),
        ("transactions", "description", "varchar", "NULL", "Free-text note."),
        ("transactions", "transfer_id", "uuid", "FK transfers.id, NULL", "Set when the entry is a transfer leg."),
        ("transactions", "created_at", "timestamptz", "default now()", "Entry time."),
        # transfers
        ("transfers", "id", "uuid", "PK", "Primary key."),
        ("transfers", "from_wallet_id", "uuid", "FK wallets.id", "Source wallet."),
        ("transfers", "to_wallet_id", "uuid", "FK wallets.id", "Destination wallet."),
        ("transfers", "amount", "decimal(20,4)", "NOT NULL", "Transfer amount."),
        ("transfers", "currency", "varchar(3)", "NOT NULL", "Currency (must match both wallets)."),
        ("transfers", "status", "enum", "default COMPLETED", "PENDING | COMPLETED | FAILED | REVERSED."),
        ("transfers", "reference", "varchar", "UNIQUE, NULL", "Idempotency key (optional)."),
        ("transfers", "description", "varchar", "NULL", "Free-text note."),
        ("transfers", "created_at", "timestamptz", "default now()", "Creation time."),
        ("transfers", "completed_at", "timestamptz", "NULL", "Set when the transfer completes."),
    ],
    "vault": [
        # documents
        ("documents", "id", "uuid", "PK", "Primary key."),
        ("documents", "user_id", "uuid", "indexed", "Owner — plain reference to auth users.id (no cross-DB FK)."),
        ("documents", "filename", "varchar", "NOT NULL", "Original upload filename."),
        ("documents", "storage_key", "varchar", "UNIQUE, NOT NULL", "Object key in MinIO (holds ciphertext)."),
        ("documents", "mime_type", "varchar", "NOT NULL", "Reported content type."),
        ("documents", "size_bytes", "int", "NOT NULL", "Plaintext size in bytes."),
        ("documents", "checksum", "varchar", "NOT NULL", "sha256 of plaintext; verified on download."),
        ("documents", "category", "varchar", "NULL", "Optional user category/tag."),
        ("documents", "status", "enum", "default ACTIVE", "ACTIVE | DELETED (soft delete; blob retained)."),
        ("documents", "created_at", "timestamptz", "default now()", "Upload time."),
        ("documents", "updated_at", "timestamptz", "@updatedAt", "Last update time."),
        ("documents", "deleted_at", "timestamptz", "NULL", "Soft-delete marker."),
        # document_shares
        ("document_shares", "id", "uuid", "PK", "Primary key."),
        ("document_shares", "document_id", "uuid", "FK documents.id", "Shared document."),
        ("document_shares", "token_hash", "varchar", "UNIQUE, NOT NULL", "sha256 of the share token."),
        ("document_shares", "created_by", "uuid", "NOT NULL", "User who created the share."),
        ("document_shares", "expires_at", "timestamptz", "NULL", "Optional expiry."),
        ("document_shares", "max_downloads", "int", "NULL", "Optional download cap."),
        ("document_shares", "download_count", "int", "default 0", "Downloads used so far."),
        ("document_shares", "revoked_at", "timestamptz", "NULL", "Set when revoked."),
        ("document_shares", "created_at", "timestamptz", "default now()", "Creation time."),
    ],
}

# ---------------------------------------------------------------------------
# STYLING
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SERVICE_FILL = PatternFill("solid", fgColor="E5EDFF")
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build():
    wb = Workbook()

    # ----- Tables overview sheet -----
    ws = wb.active
    ws.title = "Tables"
    ws["A1"] = "VaultPay — Table Catalog (all services)"
    ws["A1"].font = TITLE_FONT
    ws.append([])
    headers = ["Service", "Table", "Purpose", "Reason / Why", "Description"]
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))
    for svc, table, purpose, reason, desc in TABLES:
        ws.append([svc, table, purpose, reason, desc])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws.cell(row=ws.max_row, column=1).fill = SERVICE_FILL
    autosize(ws, [12, 22, 38, 50, 50])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ----- per-service column sheets -----
    for svc, rows in COLUMNS.items():
        s = wb.create_sheet(title=f"{svc}-columns")
        s["A1"] = f"{svc}-service — column dictionary"
        s["A1"].font = TITLE_FONT
        s.append([])
        h = ["Table", "Column", "Type", "Constraints", "Description"]
        s.append(h)
        hr = s.max_row
        style_header(s, hr, len(h))
        prev = None
        for table, col, typ, cons, desc in rows:
            s.append([table, col, typ, cons, desc])
            for c in range(1, len(h) + 1):
                cell = s.cell(row=s.max_row, column=c)
                cell.alignment = WRAP_TOP
                cell.border = BORDER
            if table != prev:
                s.cell(row=s.max_row, column=1).fill = SERVICE_FILL
                s.cell(row=s.max_row, column=1).font = Font(bold=True)
            prev = table
        autosize(s, [22, 26, 14, 22, 55])
        s.freeze_panes = s.cell(row=hr + 1, column=1)

    out = os.path.join(os.path.dirname(__file__), "vaultpay-data-dictionary.xlsx")
    wb.save(out)
    print("wrote", out, "—", len(TABLES), "tables")


if __name__ == "__main__":
    build()
